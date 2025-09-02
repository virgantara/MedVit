import torch
import torch.nn as nn
import pandas as pd
import os
from openpyxl import load_workbook
from torchvision import transforms
from PIL import Image
import numpy as np
import cv2
import random

class FocalCE(nn.Module):
    def __init__(self, weight=None, gamma=2.0):
        super().__init__()
        self.weight = weight
        self.gamma = gamma
    def forward(self, logits, target):
        ce = nn.functional.cross_entropy(logits, target, weight=self.weight, reduction='none')
        pt = torch.exp(-ce)                # pt = softmax prob of the true class
        focal = ((1-pt)**self.gamma) * ce
        return focal.mean()

class CLAHE(object):
    """
    CLAHE transform for PIL Images.
    - Apply BEFORE ToTensor().
    - Works with grayscale or RGB PIL Images.
    - Use `p < 1.0` to make it a stochastic augmentation.
    """
    def __init__(self, clip_limit=2.0, tile_grid_size=(8, 8), p=1.0, mode='lab'):
        """
        Args:
            clip_limit (float): CLAHE clip limit (try 1.5–2.5 for X-ray).
            tile_grid_size (tuple): tile size (try (8,8) or (12,12)).
            p (float): probability to apply CLAHE.
            mode (str): 'lab' for RGB via LAB on L-channel, 'gray' for single-channel.
        """
        self.clip_limit = float(clip_limit)
        self.tile_grid_size = tuple(tile_grid_size)
        self.p = float(p)
        self.mode = mode

    def __call__(self, img: Image.Image) -> Image.Image:
        if random.random() > self.p:
            return img

        # Convert PIL -> numpy
        arr = np.array(img)

        # Handle grayscale vs RGB
        if arr.ndim == 2:  # grayscale (H,W)
            # Direct CLAHE on gray
            clahe = cv2.createCLAHE(clipLimit=self.clip_limit, tileGridSize=self.tile_grid_size)
            cl = clahe.apply(arr)
            out = cl
            # keep as single-channel PIL; you can replicate to 3ch later if needed
            return Image.fromarray(out)

        elif arr.ndim == 3 and arr.shape[2] == 3:
            if self.mode == 'lab':
                lab = cv2.cvtColor(arr, cv2.COLOR_RGB2LAB)
                l, a, b = cv2.split(lab)
                clahe = cv2.createCLAHE(clipLimit=self.clip_limit, tileGridSize=self.tile_grid_size)
                cl = clahe.apply(l)
                limg = cv2.merge((cl, a, b))
                out = cv2.cvtColor(limg, cv2.COLOR_LAB2RGB)
                return Image.fromarray(out)
            elif self.mode == 'gray':
                # Convert to gray, apply CLAHE, then back to 3ch
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
                clahe = cv2.createCLAHE(clipLimit=self.clip_limit, tileGridSize=self.tile_grid_size)
                cl = clahe.apply(gray)
                out = np.stack([cl, cl, cl], axis=2)
                return Image.fromarray(out)
            else:
                return img  # unknown mode → no-op
        else:
            # Unexpected shape → no-op
            return img

def top_k_accuracy(output, target, k=1):
    """Compute the top-k accuracy, but ensure k ≤ number of classes"""
    with torch.no_grad():
        max_k = min(k, output.size(1))  # Prevent k > number of classes
        _, pred = output.topk(max_k, dim=1, largest=True, sorted=True)
        correct = pred.eq(target.view(-1, 1).expand_as(pred))
        return correct[:, :max_k].reshape(-1).float().sum(0).item()


def append_row_to_excel(path, row_dict, sheet_name="Sheet1"):
    """
    Appends a single row (dict) to an xlsx file.
    - Creates file with header if it doesn't exist.
    - If exists, appends below the last row.
    """
    df_row = pd.DataFrame([row_dict])

    if not os.path.exists(path):
        # New file: write with header
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df_row.to_excel(writer, index=False, sheet_name=sheet_name)
        return

    # File exists: append without duplicating header
    book = load_workbook(path)
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        writer.book = book
        if sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            startrow = ws.max_row
        else:
            # create new sheet with header
            df_row.to_excel(writer, index=False, sheet_name=sheet_name)
            return
        df_row.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=startrow)
